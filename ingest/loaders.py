"""
Historical backfill loaders.

Context: Jeff Sackmann's tennis_atp / tennis_wta repos are no longer available
on GitHub (as of Aug 2026 his account lists only tennis_MatchChartingProject).
If you have a local clone from before, use `load_sackmann_dir` -- the data is
static so an old copy is still perfectly good. If not, tennis-data.co.uk is the
free fallback: match-level results back to 2000, with closing odds, which you
need anyway as a calibration baseline.

`generate_synthetic` exists so the whole pipeline is runnable end-to-end with
zero external data. Use it to verify the code works before spending money on an
API key.
"""

from __future__ import annotations

import csv
import glob
import os
import random
import sqlite3
from datetime import date, timedelta

from .db import normalize_name, now_iso, upsert_matches, upsert_ranking

# ------------------------------------------------------------------ helpers --

SERIES_TO_LEVEL = {
    "grand slam": "G", "masters 1000": "M", "masters": "M", "masters cup": "F",
    "atp500": "A", "atp250": "A", "international": "A", "international gold": "A",
    "premier": "A", "wta250": "A", "wta500": "A", "wta1000": "M",
    "challenger": "C", "itf": "S",
}


def _level(series: str | None, tournament: str | None = None) -> str:
    s = (series or "").strip().lower()
    if s in SERIES_TO_LEVEL:
        return SERIES_TO_LEVEL[s]
    t = (tournament or "").lower()
    if "challenger" in t:
        return "C"
    return "A"


def _int_or_none(v) -> int | None:
    try:
        i = int(float(str(v).strip()))
        return i if i > 0 else None
    except (ValueError, TypeError):
        return None


def _float_or_none(v) -> float | None:
    try:
        f = float(str(v).strip())
        return f if f > 1.0 else None
    except (ValueError, TypeError):
        return None


def ensure_player(conn: sqlite3.Connection, name: str, tour: str,
                  dob: str | None = None, country: str | None = None) -> str:
    """
    Get-or-create a canonical player by normalized name.

    Deterministic id derived from the normalized name so re-runs are stable.
    Real Sackmann ids are better if you have them -- pass them through instead.
    """
    norm = normalize_name(name)
    if not norm:
        return ""
    row = conn.execute(
        "SELECT player_id FROM players WHERE norm_name = ? AND tour = ?", (norm, tour)
    ).fetchone()
    if row:
        return row["player_id"]
    pid = f"{tour.lower()}_{norm.replace(' ', '_')}"[:60]
    conn.execute(
        """INSERT OR IGNORE INTO players (player_id, full_name, norm_name, tour, dob, country)
           VALUES (?,?,?,?,?,?)""",
        (pid, name, norm, tour, dob, country),
    )
    return pid


# ------------------------------------------------- tennis-data.co.uk loader --

# Their files are one .xlsx per year: e.g. 2025/2025.xlsx (ATP),
# w2025/w2025.xlsx (WTA). Download manually, then point this at the folder.
# VERIFY the header row against a real file before trusting the mapping below --
# they have changed column names between seasons.
TD_COLUMNS = {
    "date": "Date", "tournament": "Tournament", "series": "Series",
    "surface": "Surface", "round": "Round", "winner": "Winner", "loser": "Loser",
    "wrank": "WRank", "lrank": "LRank", "comment": "Comment",
    # closing odds -- Pinnacle is the sharpest, use it as the benchmark
    "psw": "PSW", "psl": "PSL", "avgw": "AvgW", "avgl": "AvgL",
}


def load_tennis_data_file(conn: sqlite3.Connection, path: str, tour: str = "ATP") -> dict:
    """Load one tennis-data.co.uk .xlsx or .csv into the match store."""
    rows_raw = _read_table(path)
    if not rows_raw:
        return {"seen": 0, "inserted": 0, "updated": 0}

    missing = [k for k in ("date", "winner", "loser") if TD_COLUMNS[k] not in rows_raw[0]]
    if missing:
        raise ValueError(
            f"{path}: expected columns {missing} not found. "
            f"Got: {list(rows_raw[0])[:15]}. Fix TD_COLUMNS."
        )

    out, odds = [], []
    for r in rows_raw:
        w, l = r.get(TD_COLUMNS["winner"]), r.get(TD_COLUMNS["loser"])
        if not w or not l:
            continue
        d = _parse_date(r.get(TD_COLUMNS["date"]))
        if not d:
            continue

        comment = str(r.get(TD_COLUMNS["comment"], "") or "")
        tournament = r.get(TD_COLUMNS["tournament"])
        surface = r.get(TD_COLUMNS["surface"])

        wid = ensure_player(conn, str(w), tour)
        lid = ensure_player(conn, str(l), tour)

        out.append({
            "match_date": d,
            "tour": tour,
            "tour_level": _level(r.get(TD_COLUMNS["series"]), str(tournament)),
            "tournament": str(tournament) if tournament else None,
            "round": str(r.get(TD_COLUMNS["round"]) or "") or None,
            "surface": str(surface).strip().title() if surface else None,
            "winner_id": wid, "loser_id": lid,
            "winner_name": str(w), "loser_name": str(l),
            "winner_rank": _int_or_none(r.get(TD_COLUMNS["wrank"])),
            "loser_rank": _int_or_none(r.get(TD_COLUMNS["lrank"])),
            "score": None,
            "retirement": "ret" in comment.lower() or "w/o" in comment.lower(),
            "source_event_key": None,
        })
        odds.append((
            _float_or_none(r.get(TD_COLUMNS["psw"])) or _float_or_none(r.get(TD_COLUMNS["avgw"])),
            _float_or_none(r.get(TD_COLUMNS["psl"])) or _float_or_none(r.get(TD_COLUMNS["avgl"])),
        ))

    conn.commit()
    counts = upsert_matches(conn, out, "tennis-data")
    _store_odds(conn, out, odds)
    return counts


def _store_odds(conn: sqlite3.Connection, matches: list[dict], odds: list[tuple]) -> None:
    """
    Closing odds are the calibration benchmark -- NOT a training feature.

    Training on them teaches the model to imitate the market, which guarantees
    you never beat it. Kept in a separate table so it's hard to leak by accident.
    """
    conn.execute("""
        CREATE TABLE IF NOT EXISTS closing_odds (
            match_key TEXT PRIMARY KEY, winner_odds REAL, loser_odds REAL,
            ingested_at TEXT
        )""")
    from .db import make_match_key
    for m, (wo, lo) in zip(matches, odds):
        if wo is None or lo is None:
            continue
        key = make_match_key(m["match_date"], m.get("tournament", ""),
                             m["winner_name"], m["loser_name"])
        conn.execute(
            """INSERT INTO closing_odds VALUES (?,?,?,?)
               ON CONFLICT(match_key) DO UPDATE SET
                   winner_odds=excluded.winner_odds, loser_odds=excluded.loser_odds""",
            (key, wo, lo, now_iso()),
        )
    conn.commit()


def load_tennis_data_dir(conn: sqlite3.Connection, folder: str) -> dict:
    """Load every .xlsx/.csv in a folder. WTA files are prefixed 'w'."""
    total = {"seen": 0, "inserted": 0, "updated": 0}
    files = sorted(glob.glob(os.path.join(folder, "*.xlsx")) +
                   glob.glob(os.path.join(folder, "*.csv")))
    for path in files:
        tour = "WTA" if os.path.basename(path).lower().startswith("w") else "ATP"
        try:
            c = load_tennis_data_file(conn, path, tour)
            print(f"  {os.path.basename(path):20s} {tour}  +{c['inserted']}")
            for k in total:
                total[k] += c[k]
        except Exception as exc:
            print(f"  {os.path.basename(path):20s} SKIPPED: {exc}")
    return total


def _read_table(path: str) -> list[dict]:
    if path.lower().endswith((".xlsx", ".xls")):
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows)]
        return [dict(zip(header, r)) for r in rows if any(v is not None for v in r)]
    with open(path, newline="", encoding="utf-8-sig", errors="replace") as f:
        return list(csv.DictReader(f))


def _parse_date(v) -> str | None:
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        return v.isoformat()[:10]
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d/%m/%y"):
        try:
            from datetime import datetime
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


# ------------------------------------------------------- Sackmann (if local) --

SACKMANN_SURFACE = {"Hard", "Clay", "Grass", "Carpet"}


def load_sackmann_dir(conn: sqlite3.Connection, folder: str, tour: str = "ATP") -> dict:
    """
    Load an existing local clone of tennis_atp / tennis_wta.

    The repos are gone from GitHub, but the data is static -- an old clone is
    still fully valid. Reads atp_matches_*.csv and atp_players.csv.
    """
    total = {"seen": 0, "inserted": 0, "updated": 0}

    pfile = glob.glob(os.path.join(folder, "*_players.csv"))
    if pfile:
        with open(pfile[0], newline="", encoding="utf-8", errors="replace") as f:
            for r in csv.DictReader(f):
                pid = r.get("player_id")
                full = f"{r.get('name_first','')} {r.get('name_last','')}".strip()
                if not pid or not full:
                    continue
                dob = r.get("dob") or ""
                dob_iso = f"{dob[:4]}-{dob[4:6]}-{dob[6:8]}" if len(dob) == 8 else None
                conn.execute(
                    """INSERT OR IGNORE INTO players
                       (player_id, full_name, norm_name, tour, dob, country)
                       VALUES (?,?,?,?,?,?)""",
                    (pid, full, normalize_name(full), tour, dob_iso, r.get("ioc")),
                )
        conn.commit()
        print(f"  players loaded from {os.path.basename(pfile[0])}")

    for path in sorted(glob.glob(os.path.join(folder, "*_matches_*.csv"))):
        if "qual" in os.path.basename(path) and "chall" not in os.path.basename(path):
            pass  # qualifying/challenger files are fine to include
        rows = []
        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            for r in csv.DictReader(f):
                d = r.get("tourney_date", "")
                if len(d) != 8:
                    continue
                rows.append({
                    "match_date": f"{d[:4]}-{d[4:6]}-{d[6:8]}",
                    "tour": tour,
                    "tour_level": r.get("tourney_level"),
                    "tournament": r.get("tourney_name"),
                    "round": r.get("round"),
                    "surface": r.get("surface") if r.get("surface") in SACKMANN_SURFACE else None,
                    "winner_id": r.get("winner_id"),
                    "loser_id": r.get("loser_id"),
                    "winner_name": r.get("winner_name", ""),
                    "loser_name": r.get("loser_name", ""),
                    "winner_rank": _int_or_none(r.get("winner_rank")),
                    "loser_rank": _int_or_none(r.get("loser_rank")),
                    "score": r.get("score"),
                    "retirement": "RET" in str(r.get("score", "")).upper(),
                    "source_event_key": None,
                })
        c = upsert_matches(conn, rows, "sackmann")
        print(f"  {os.path.basename(path):28s} +{c['inserted']}")
        for k in total:
            total[k] += c[k]

    for path in sorted(glob.glob(os.path.join(folder, "*_rankings_*.csv"))):
        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            for r in csv.DictReader(f):
                d = r.get("ranking_date", "")
                pid, rank = r.get("player"), _int_or_none(r.get("rank"))
                if len(d) != 8 or not pid or not rank:
                    continue
                upsert_ranking(conn, pid, tour, rank, _int_or_none(r.get("points")),
                               f"{d[:4]}-{d[4:6]}-{d[6:8]}")
        conn.commit()
        print(f"  rankings from {os.path.basename(path)}")

    return total


# ------------------------------------------------------------- synthetic ----

def generate_synthetic(conn: sqlite3.Connection, n_players: int = 120,
                       n_matches: int = 4000, seed: int = 7) -> dict:
    """
    Generate a realistic-shaped fake dataset so the pipeline runs end-to-end
    with no external data.

    Each player gets a latent skill; match outcomes follow a logistic function
    of the skill gap plus a surface-specific bonus. That means a well-built
    model SHOULD recover signal here -- if it can't beat 50% on synthetic data,
    the bug is in your code, not your features.
    """
    rng = random.Random(seed)
    surfaces = ["Hard", "Clay", "Grass"]
    levels = ["G", "M", "A", "C"]

    # Distinct surname-like names. NOTE: normalize_name() strips digits and
    # single letters, so names like 'Player A001' would ALL collapse to
    # 'player' and resolve to one id -- which silently turns every match into
    # a player-vs-himself row with zero-variance features. Real names only.
    syl_a = ["Bar", "Kov", "Mar", "Del", "Fer", "Nov", "Ras", "Tor", "Vel", "Zan",
             "Cor", "Hal", "Jur", "Lek", "Mun", "Pav", "Rud", "Sav", "Tam", "Vin"]
    syl_b = ["ic", "ov", "en", "ez", "ini", "sky", "aud", "ard", "ell", "ossi"]
    names, used = [], set()
    for a in syl_a:
        for b in syl_b:
            n = f"{a}{b}"
            if n.lower() not in used:
                used.add(n.lower())
                names.append(n)
    rng.shuffle(names)
    if n_players > len(names):
        raise ValueError(f"generator supports at most {len(names)} players")

    players = []
    for i in range(n_players):
        name = f"{names[i]} {chr(65 + i % 26)}."
        pid = ensure_player(conn, name, "ATP")
        players.append({
            "id": pid,
            "skill": rng.gauss(0, 1),
            "surf": {s: rng.gauss(0, 0.35) for s in surfaces},
        })
    conn.commit()

    # Guard: distinct names MUST yield distinct ids. If this trips, the name
    # generator is colliding under normalization.
    ids = {p["id"] for p in players}
    if len(ids) != n_players:
        raise RuntimeError(
            f"player id collision: {n_players} names -> {len(ids)} ids. "
            "Names are colliding under normalize_name()."
        )

    # rankings derived from latent skill, snapshotted monthly
    ordered = sorted(players, key=lambda p: -p["skill"])
    start = date.today() - timedelta(days=730)
    for month in range(0, 25):
        snap = (start + timedelta(days=30 * month)).isoformat()
        for rank, p in enumerate(ordered, start=1):
            jitter = rng.randint(-4, 4)
            upsert_ranking(conn, p["id"], "ATP", max(1, rank + jitter),
                           2000 - rank * 12, snap)
    conn.commit()

    rows, odds = [], []
    for m in range(n_matches):
        d = start + timedelta(days=rng.randint(20, 725))
        surface = rng.choice(surfaces)
        a, b = rng.sample(players, 2)
        ea = a["skill"] + a["surf"][surface]
        eb = b["skill"] + b["surf"][surface]
        p_a = 1 / (1 + pow(2.718281828, -(ea - eb)))
        a_wins = rng.random() < p_a
        w, l = (a, b) if a_wins else (b, a)

        wname = conn.execute("SELECT full_name FROM players WHERE player_id=?",
                             (w["id"],)).fetchone()["full_name"]
        lname = conn.execute("SELECT full_name FROM players WHERE player_id=?",
                             (l["id"],)).fetchone()["full_name"]
        tournament = f"Synthetic Event {m % 40}"

        rows.append({
            "match_date": d.isoformat(), "tour": "ATP",
            "tour_level": rng.choice(levels), "tournament": tournament,
            "round": rng.choice(["R32", "R16", "QF", "SF", "F"]), "surface": surface,
            "winner_id": w["id"], "loser_id": l["id"],
            "winner_name": wname, "loser_name": lname,
            "winner_rank": None, "loser_rank": None,
            "score": "6-4 6-3", "retirement": rng.random() < 0.03,
            "source_event_key": f"syn_{m}",
        })
        # market odds = true probability + noise + 5% vig
        p_true = p_a if a_wins else (1 - p_a)
        p_mkt = min(0.97, max(0.03, p_true + rng.gauss(0, 0.04)))
        odds.append((round(1 / (p_mkt * 1.05), 3), round(1 / ((1 - p_mkt) * 1.05), 3)))

    # backfill point-in-time ranks
    from .db import rank_as_of
    for r in rows:
        r["winner_rank"] = rank_as_of(conn, r["winner_id"], r["match_date"])
        r["loser_rank"] = rank_as_of(conn, r["loser_id"], r["match_date"])

    counts = upsert_matches(conn, rows, "synthetic")
    _store_odds(conn, rows, odds)
    return counts
